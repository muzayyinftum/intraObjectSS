import sklearn.metrics as sm
import openpyxl as xl
import scipy.io.wavfile
import math
import numpy as np
import os
import shutil
import importlib

methods = importlib.import_module("2-METHOD")

def sampling(file_audio, return_rate=False):
    rate, data = scipy.io.wavfile.read(file_audio)
    data = np.array(data,dtype=np.int16)
    data = np.add(data,[32768])
    if return_rate:
        return rate, data
    return data

def mean_data_sample(data_sample):
    return np.mean(np.power(data_sample,[2]))

def calculate_mse(data_sample, data_stego):
    return sm.mean_squared_error(data_sample,data_stego)

def calculate_snr(data_sample,mse):
    if mse == 0:
        return 'infinite'
    else:
        mds = mean_data_sample(data_sample)
        log_content = mds/mse
        return 10 * math.log(log_content,10)

def calculate_psnr(mse):
    if mse == 0:
        return 'infinite'
    else:
        log_content = (((2 ** 16) - 1) ** 2)/mse
        return 10 * math.log(log_content,10)

def clone_cover_audio(data_sample, filename, frame_rate, current_L):
    interpolated_sample = methods.catmull_rom_interpolation(data_sample, current_L)
    new_data = methods.combine(interpolated_sample, data_sample, current_L)
    process_data = np.subtract(new_data,[32768])
    process_data = np.array(process_data,dtype=np.int16)
    os.makedirs(os.path.dirname(filename), exist_ok=True)
    scipy.io.wavfile.write(filename, frame_rate * (current_L + 1), process_data)
    return new_data

def clone_cover_audio_file(file_audio, filename, current_L):
    frame_rate, data_sample = sampling(file_audio, return_rate=True)
    return clone_cover_audio(data_sample, filename, frame_rate, current_L)

def get_interpolation_level(original_frame_rate, stego_frame_rate):
    if stego_frame_rate % original_frame_rate != 0:
        raise ValueError("Frame rate stego tidak sesuai dengan frame rate original.")

    current_L = (stego_frame_rate // original_frame_rate) - 1
    if current_L < 1:
        raise ValueError("Nilai L dari frame rate stego tidak valid.")

    return current_L

def getAvg(data_psnr):
    new = np.transpose(data_psnr)

    data_avg = []
    for x in range(len(new)):
        data_avg.append(np.mean(new[x]))
    return data_avg

def print_excel(data_mse, data_snr, data_psnr, filename):
    excel = xl.Workbook()
    sheet_mse = excel.create_sheet('Mean Squared Error')

    total_audio = len(data_mse)
    total_payload = len(data_mse[0])

    for index_audio in range (0,total_audio):
        if index_audio == 0:
            for x in range (0, total_payload):
                sheet_mse.cell(row=1,column=x+2).value = 'Payload'+str(x+1)

        for index_payload in range (0,total_payload):
            if index_payload == 0:
                sheet_mse.cell(row=index_audio+2,column=1).value = 'Audio'+str(index_audio+1)
            sheet_mse.cell(row = index_audio+2, column=index_payload+2).value = data_mse[index_audio][index_payload]

    sheet_snr = excel.create_sheet('SNR')

    total_audio = len(data_snr)
    total_payload = len(data_snr[0])

    for index_audio in range(0, total_audio):
        if index_audio == 0:
            for x in range (0, total_payload):
                sheet_snr.cell(row=1,column=x+2).value = 'Payload'+str(x+1)

        for index_payload in range(0, total_payload):
            if index_payload == 0:
                sheet_snr.cell(row=index_audio + 2, column=1).value = 'Audio' + str(index_audio + 1)
            sheet_snr.cell(row=index_audio + 2, column=index_payload + 2).value = data_snr[index_audio][index_payload]

    sheet_psnr = excel.create_sheet('PSNR')

    total_audio = len(data_psnr)
    total_payload = len(data_psnr[0])

    for index_audio in range(0, total_audio):
        if index_audio == 0:
            for x in range (0, total_payload):
                sheet_psnr.cell(row=1,column=x+2).value = 'Payload'+str(x+1)

        for index_payload in range(0, total_payload):
            if index_payload == 0:
                sheet_psnr.cell(row=index_audio + 2, column=1).value = 'Audio' + str(index_audio + 1)
            sheet_psnr.cell(row=index_audio + 2, column=index_payload + 2).value = data_psnr[index_audio][index_payload]

    excel.save(filename)

def main():

    folder_sample_audio = 'stegoaudioDataset/Audio/'
    folder_output_clone = 'audio_clone/'

    file_audio = [folder_sample_audio+'data'+str(x)+'_mono.wav' for x in range (1,16)]

    file_stego_audio = []
    for index_audio in range (1,16):
        file_stego_audio.append(['embeddingResults/stego_audio' + str(index_audio) + '_payload' + str(index_payload) + '.wav' for index_payload in range(1, 12)])


    data_mse = []
    data_snr = []
    data_psnr = []
    for index_audio in range (0,15):
        mse = []
        snr = []
        psnr = []
        for index_payload in range (0,11):
            original_frame_rate, original_sample = sampling(file_audio[index_audio], return_rate=True)
            stego_frame_rate, sample_stego_audio = sampling(file_stego_audio[index_audio][index_payload], return_rate=True)
            current_L = get_interpolation_level(original_frame_rate, stego_frame_rate)
            clone_filename = folder_output_clone+'audio'+str(index_audio+1)+'_payload'+str(index_payload+1)+'mono.wav'
            sample_audio = clone_cover_audio(original_sample, clone_filename, original_frame_rate, current_L)

            mse.append(calculate_mse(sample_audio,sample_stego_audio))
            snr.append(calculate_snr(sample_audio,mse[index_payload]))
            psnr.append(calculate_psnr(mse[index_payload]))
        data_mse.append(mse)
        data_snr.append(snr)
        data_psnr.append(psnr)


    filename = 'quality_result.xlsx'
    print_excel(data_mse, data_snr, data_psnr, filename)

if __name__ == '__main__':
    main()
